# spectraconverter_v4/src/data_exporter.py

import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Alignment, Font

def export_to_excel(spectra_data, output_path):
    """
    Exporta los datos crudos y procesados a un archivo Excel con un gráfico.
    """
    wb = Workbook()
    
    ws_proc = wb.active
    ws_proc.title = "Datos Procesados"
    
    col_index = 1
    for spec in spectra_data:
        df_to_write = spec['processed_dataframe'] if spec['processed_dataframe'] is not None else spec['dataframe']
        ws_proc.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + 1)
        title_cell = ws_proc.cell(row=1, column=col_index, value=spec['filename'])
        title_cell.font = Font(bold=True, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for r in dataframe_to_rows(df_to_write, index=False, header=True):
            ws_proc.append(r)
        
        last_row = ws_proc.max_row
        first_row = last_row - len(df_to_write)
        ws_proc.move_range(f"A{first_row}:{ws_proc.cell(row=last_row, column=2).column_letter}{last_row}", 
                           rows=-(first_row - 2), cols=col_index - 1)

        col_index += 3

    ws_raw = wb.create_sheet("Datos Crudos")
    col_index = 1
    for spec in spectra_data:
        df_to_write = spec['dataframe']
        ws_raw.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + 1)
        title_cell = ws_raw.cell(row=1, column=col_index, value=spec['filename'])
        title_cell.font = Font(bold=True, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for r in dataframe_to_rows(df_to_write, index=False, header=True):
            ws_raw.append(r)
        last_row = ws_raw.max_row
        first_row = last_row - len(df_to_write)
        ws_raw.move_range(f"A{first_row}:{ws_raw.cell(row=last_row, column=2).column_letter}{last_row}", 
                           rows=-(first_row - 2), cols=col_index - 1)
        col_index += 3

    chart = ScatterChart()
    chart.title = "Espectros Procesados"
    chart.x_axis.title = "Longitud de onda (nm)"
    chart.y_axis.title = "Intensidad / Absorbancia"
    chart.legend.position = 'b'

    for i, spec in enumerate(spectra_data):
        col = i * 3 + 1
        x_ref = Reference(ws_proc, min_col=col, min_row=3, max_row=ws_proc.max_row)
        y_ref = Reference(ws_proc, min_col=col + 1, min_row=2, max_row=ws_proc.max_row)
        series = Series(y_ref, x_ref, title_from_data=True)
        chart.series.append(series)

    chart_anchor_col = ws_proc.cell(row=1, column=col_index + 1).column_letter
    ws_proc.add_chart(chart, f"{chart_anchor_col}2")
    
    wb.save(output_path)


def export_to_scidavis(spectra_data, output_path):
    """
    Exporta los datos procesados a un archivo .tsv (tab-separated) para SciDAVis/Origin.
    """
    if not spectra_data:
        return

    def sanitize_header(fname):
        """Limpia un nombre de archivo para que sea una cabecera de columna segura."""
        name_without_ext = os.path.splitext(fname)[0]
        sanitized_name = re.sub(r'[^a-zA-Z0-9_]', '_', name_without_ext)
        return sanitized_name

    base_spec = spectra_data[0]
    df_final = base_spec['processed_dataframe'] if base_spec['processed_dataframe'] is not None else base_spec['dataframe']
    df_final = df_final.rename(columns={'intensity': sanitize_header(base_spec['filename'])})

    for spec in spectra_data[1:]:
        df_to_merge = spec['processed_dataframe'] if spec['processed_dataframe'] is not None else spec['dataframe']
        df_to_merge = df_to_merge.rename(columns={'intensity': sanitize_header(spec['filename'])})
        df_final = pd.merge(df_final, df_to_merge, on='wavelength', how='outer')

    df_final = df_final.sort_values(by='wavelength').reset_index(drop=True)
    df_final = df_final.interpolate(method='linear', limit_direction='both', axis=0)
    df_final.fillna(0, inplace=True)
    
    # Exportamos con coma decimal, que es lo más compatible para importación manual en sistemas en español.
    df_final.to_csv(
        output_path, 
        sep='\t',
        index=False,
        decimal=',',
        float_format='%.6f',
        encoding='utf-8-sig'
    )